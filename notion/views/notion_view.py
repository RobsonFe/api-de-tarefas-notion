from asyncio.log import logger
from notion.serializer.serializers import NotionSerializer
from notion.models.entity.notion import Notion
from rest_framework.response import Response
from rest_framework import generics
from rest_framework import status
from dotenv import load_dotenv
import logging
import openpyxl
import requests
import json
import os

load_dotenv()

# Defina o nível de logging como INFO ou DEBUG
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

notion_token = os.getenv("NOTION_TOKEN")
headers = {
    'Authorization': f"Bearer {notion_token}",
    'Content-Type': 'application/json',
    'Notion-Version': '2022-02-22'
}
banco_notion = os.getenv("ID_DO_BANCO")

if not notion_token or not banco_notion:
    raise ValueError(
        "Token de acesso ao Notion ou ID do banco não encontrado no arquivo .env"
    )

# Buscar dados no Notion


def get_data_from_notion():
    url = f"https://api.notion.com/v1/database/{banco_notion}/query"
    payload = {"page_size": 100}
    response = requests.post(url, json=payload, headers=headers)
    data = response.json()
    return data['results']

# Deletar dados do Notion


def delete_from_notion(notion_page_id):
    url = f"https://api.notion.com/v1/pages/{notion_page_id}"
    headers = {
        "Authorization": f"Bearer {notion_token}",
        "Content-Type": "application/json",
        "Notion-Version": "2022-06-28",
    }
    data = {
        "archived": True
    }
    response = requests.patch(url, json=data, headers=headers)
    if response.status_code != 200:
        logger.error("Erro ao excluir a página no Notion: %s", response.text)
        raise Exception("Erro ao excluir a página no Notion")
    logger.info(
        "Página com Notion Page ID %s arquivada no Notion.", notion_page_id)

# Excluir dados do Excel


def delete_from_sheet(notion_page_id):
    file_path = "./planilhas/Tarefas.xlsx"
    sheet_name = "Tarefas"

    if not os.path.exists(file_path):
        logger.warning("O arquivo de planilha não foi encontrado.")
        return

    workbook = openpyxl.load_workbook(file_path)

    if sheet_name not in workbook.sheetnames:
        logger.warning("A planilha '%s' não existe no arquivo.", sheet_name)
        return

    worksheet = workbook[sheet_name]

    # Encontre e exclua a linha com o Notion Page ID correspondente
    rows_to_delete = []
    for row in worksheet.iter_rows(min_row=2, max_col=4):
        if row[3].value == notion_page_id:
            rows_to_delete.append(row[0].row)

    for row_idx in reversed(rows_to_delete):
        worksheet.delete_rows(row_idx)

    workbook.save(file_path)
    logger.info(f"Linhas com Notion Page ID {
                notion_page_id} excluídas da planilha.")


# Excluir do banco de dados

def delete_from_db(notion_page_id):
    try:
        Notion.objects.get(notion_page_id=notion_page_id).delete()
        logger.info(
            "Registro com Notion Page ID %s excluído do banco de dados.", notion_page_id)
    except Notion.DoesNotExist:
        logger.warning(
            "Registro com Notion Page ID %s não encontrado no banco de dados.", notion_page_id)


# Função para salvar ou atualizar dados em uma planilha de Excel


def save_or_update_in_sheet(notion_data):
    file_path = "./planilhas/Tarefas.xlsx"
    sheet_name = "Tarefas"

    # Tente abrir o arquivo, ou crie um novo workbook se não existir
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
    else:
        workbook = openpyxl.Workbook()

    # Verifique se a worksheet já existe, senão crie uma nova
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.active
        worksheet.title = sheet_name
        # Cabeçalho correto
        worksheet.append(["Title", "Status", "Priority", "Notion Page ID"])

    # Verifique se o ID já existe para atualizar a linha correspondente
    id_exists = False
    for row in worksheet.iter_rows(min_row=2, values_only=False):
        if row[3].value == notion_data["notion_page_id"]:  # Notion Page ID está na 4ª coluna
            # Atualize a linha existente
            row[0].value = notion_data["title"]
            row[1].value = notion_data["status"]
            row[2].value = notion_data["priority"]
            id_exists = True
            break

    # Se o ID não existir, adicione uma nova linha
    if not id_exists:
        worksheet.append([
            notion_data["title"],              # Title na 1ª coluna
            notion_data["status"],             # Status na 2ª coluna
            notion_data["priority"],           # Priority na 3ª coluna
            notion_data["notion_page_id"]     # Notion Page ID na 4ª coluna
        ])

    # Salve a planilha
    workbook.save(file_path)
    logger.info(f"Planilha atualizada e salva em: {file_path}")

# Criar Dados


class NotionCreateView(generics.CreateAPIView):
    serializer_class = NotionSerializer

    def create(self, request, *args, **kwargs):
        try:
            data = request.data
            url = "https://api.notion.com/v1/pages"
            payload = {
                "parent": {"database_id": banco_notion},
                "properties": {
                    "title": {"title": [{"text": {"content": data["title"]}}]},
                    "Prioridade": {"select": {"name": data["priority"]}},
                    "Status": {"status": {"name": data["status"]}},
                }
            }
            response = requests.post(url, json=payload, headers=headers)

            # Verificação se a resposta foi bem-sucedida
            if response.status_code in [200, 201]:
                # Salvar a tarefa no banco de dados
                notion_id = response.json()["id"]
                notion = Notion.objects.create(
                    title=data["title"],
                    status=data["status"],
                    priority=data["priority"],
                    notion_page_id=notion_id
                )

                # Serializar os dados do objeto criado para retorno das respostas
                serialized_notion = NotionSerializer(notion).data

                # Atualizar a planilha
                notion_data = {
                    "notion_page_id": notion_id,
                    "title": data["title"],
                    "status": data["status"],
                    "priority": data["priority"]
                }
                save_or_update_in_sheet(notion_data)

                # Registrar os dados no log em formato JSON
                logger.info(json.dumps(serialized_notion,
                            indent=4, ensure_ascii=False))

            # Retornar os dados criados na resposta
            return Response({"message": "Tarefa criada com sucesso", "data": serialized_notion}, status=status.HTTP_201_CREATED)
        except Exception as erro:
            logger.error("Erro ao criar Notion: %s", erro)
            return Response({"message": "Erro ao criar tarefa"}, status=status.HTTP_400_BAD_REQUEST)

# Atualizar e Deletar Dados


class NotionUpdateAndDelete(generics.RetrieveUpdateDestroyAPIView):
    queryset = Notion.objects.all()
    serializer_class = NotionSerializer
    lookup_field = 'pk'

    def update(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            data = request.data

            # Atualizar no banco de dados
            serializer = self.get_serializer(instance, data=data, partial=True)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            updated_notion = serializer.instance

            # Atualizar a página no Notion
            notion_update_data = {
                "properties": {
                    "title": {
                        "title": [
                            {
                                "text": {
                                    "content": data.get("title", updated_notion.title),
                                },
                            },
                        ],
                    },
                    "Prioridade": {
                        "select": {
                            "name": data.get("priority", updated_notion.priority),
                        },
                    },
                    "Status": {
                        "status": {
                            "name": data.get("status", updated_notion.status),
                        },
                    },
                },
            }

            url = f"https://api.notion.com/v1/pages/{
                updated_notion.notion_page_id}"
            response = requests.patch(
                url, json=notion_update_data, headers=headers)

            if response.status_code not in [200, 202]:
                logger.error(
                    "Erro ao atualizar a página no Notion: %s", response.text)
                raise Exception("Erro ao atualizar a página no Notion")

            # Atualizar a planilha de Excel
            notion_data = {
                "notion_page_id": updated_notion.notion_page_id,
                "title": data.get("title", updated_notion.title),
                "status": data.get("status", updated_notion.status),
                "priority": data.get("priority", updated_notion.priority)
            }
            save_or_update_in_sheet(notion_data)

            # Serializar o objeto atualizado e os dados da resposta
            serialized_notion = NotionSerializer(updated_notion).data
            response_data = {
                "message": "Tarefa atualizada com sucesso", "data": serializer.data}

            # Registrar os dados no log em formato JSON
            logger.info("Dados Atualizados: %s", json.dumps(
                serialized_notion, indent=4, ensure_ascii=False))
            logger.info("Resposta: %s", json.dumps(
                response_data, indent=4, ensure_ascii=False))

            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as erro:
            logger.error("Erro ao atualizar Notion: %s", erro)
            return Response({"message": "Erro ao atualizar tarefa"}, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            notion_page_id = instance.notion_page_id

            # Obter detalhes do objeto antes da exclusão
            serialized_notion = NotionSerializer(instance).data

            # Excluir do Notion
            delete_from_notion(notion_page_id)

            # Excluir da planilha de Excel
            delete_from_sheet(notion_page_id)

            # Excluir do banco de dados
            delete_from_db(notion_page_id)
            # Registrar o objeto excluído no console
            logger.info("Objeto excluído: %s", json.dumps(
                serialized_notion, indent=4, ensure_ascii=False))

            # Retornar o objeto excluído na resposta da API
            return Response({"message": "Tarefa excluída com sucesso", "data": serialized_notion}, status=status.HTTP_204_NO_CONTENT)
        except Exception as erro:
            logger.error("Erro ao excluir Notion: %s", erro)
            return Response({"message": "Erro ao excluir tarefa"}, status=status.HTTP_400_BAD_REQUEST)

# Listar Dados


class NotionList(generics.ListCreateAPIView):
    queryset = Notion.objects.all()
    serializer_class = NotionSerializer

# Buscar dados no banco pelo ID


class NotionFindById(generics.RetrieveAPIView):
    queryset = Notion.objects.all()
    serializer_class = NotionSerializer
    lookup_field = 'pk'


class FindIdByNotion(generics.RetrieveAPIView):
    queryset = Notion.objects.all()
    serializer_class = NotionSerializer
    lookup_field = 'notion_page_id'
